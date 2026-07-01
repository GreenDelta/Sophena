#!/usr/bin/env python3
"""
Load manufacturer logos from data/Logos.zip and update manufacturers CSV.
"""
from __future__ import annotations
import base64
import os
import shutil
import csv
import zipfile
from dataclasses import dataclass
from typing import Dict, Optional
import sys
import re

csv.field_size_limit(sys.maxsize)

@dataclass
class ManufacturerData:
    data_url: str
    sort_value: float

@dataclass
class ManufacturerContact:
    address: str
    weblink: str


def load_manufacturer_logos(zip_path: str = "data/Logos.zip") -> Dict[str, ManufacturerData]:
    """
    Open `zip_path`, read png/jpg entries and return a dictionary mapping
    manufacturer name -> ManufacturerData (contains data URL and sort value).
    """
    logos: Dict[str, ManufacturerData] = {}
    if not os.path.exists(zip_path):
        return logos

    with zipfile.ZipFile(zip_path, "r") as z:
        for info in z.infolist():
            name = os.path.basename(info.filename)
            if not name:
                continue
            root, ext = os.path.splitext(name)
            ext_l = ext.lower().lstrip(".")
            if ext_l not in ("png", "jpg", "jpeg"):
                continue
            # Map extension to mime type
            mime = "image/png" if ext_l == "png" else "image/jpeg"

            # Split by first space: first part is numeric sort value
            parts = root.split(" ", 1)
            if len(parts) == 2:
                first, rest = parts
                # parse numeric sort value (int or float)
                try:
                    sort_value: float = int(first)  # try int first
                except Exception:
                    try:
                        sort_value = float(first)
                    except Exception:
                        sort_value = 0.0
            else:
                # if no space, use full root as name and default sort value
                rest = parts[0]
                sort_value = 0.0
            manufacturer = rest.strip()
            if not manufacturer:
                continue

            data = z.read(info.filename)
            b64 = base64.b64encode(data).decode("ascii")
            data_url = f"data:{mime};base64,{b64}"
            logos[manufacturer] = ManufacturerData(data_url=data_url, sort_value=sort_value)

    return logos


def load_manufacturer_contacts(excel_path: str = "data/Kontaktdaten für Software.xlsx") -> Dict[str, ManufacturerContact]:
    """
    Read the first sheet from `excel_path` and return a dictionary mapping the
    name from column A to ManufacturerContact(address=column B, weblink=column C).
    """
    contacts: Dict[str, ManufacturerContact] = {}
    if not os.path.exists(excel_path):
        return contacts

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed; skipping manufacturer contact import")
        return contacts

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        name = str(row[0]).strip() if row[0] is not None else ""
        if not name:
            continue

        address = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        weblink = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        contacts[name] = ManufacturerContact(address=address, weblink=weblink)

    return contacts


def find_best_match(name: str, logos: Dict[str, ManufacturerData]) -> Optional[ManufacturerData]:
    """Return ManufacturerData for the first logo whose normalized first word matches `name`."""
    if not name:
        return None

    def normalize(text: str) -> str:
        parts = re.split('[- ]', text.strip().lower())
        return parts[0]

    norm_name = normalize(name)
    if not norm_name:
        return None

    for key, value in logos.items():
        if normalize(key) == norm_name:
            return value

    return None


def update_manufacturers_csv(logos: Dict[str, ManufacturerData],
                             src_csv: str = "data/csv/manufacturers.csv",
                             dst_csv: str = "data/csv/manufacturers.csv",
                             contacts: Optional[Dict[str, ManufacturerContact]] = None) -> None:
    """
    Create a backup of `src_csv` as `data/csv/manufacturers.bak.csv`, read the
    backup (semicolon-separated), and write an updated `dst_csv` where for each
    line the manufacturer name is read from cell index 1 and cells 5 and 6 are
    overwritten with ManufacturerData.sort_value and ManufacturerData.data_url.
    The optional `contacts` mapping is accepted for later extensions.
    """
    if not os.path.exists(src_csv):
        print(f"Source CSV not found: {src_csv}")
        return

    if contacts is None:
        contacts = {}

    bak_path = os.path.join(os.path.dirname(src_csv), "manufacturers.bak.csv")
    shutil.copyfile(src_csv, bak_path)

    # Remove destination if exists
    if os.path.exists(dst_csv):
        os.remove(dst_csv)

    processed = 0
    matched = 0

    with open(bak_path, "r", newline="", encoding="utf-8") as infp, open(dst_csv, "w", newline="", encoding="utf-8") as outfp:
        reader = csv.reader(infp, delimiter=';')
        writer = csv.writer(outfp, delimiter=';')

        for row in reader:
            processed += 1
            # ensure we have at least 7 columns
            if len(row) < 7:
                row.extend([""] * (7 - len(row)))

            if processed == 1:
                writer.writerow(row)
                continue

            manufacturer = row[1].strip() if len(row) > 1 else ""
            mdata = find_best_match(manufacturer, logos)
            if mdata is not None:
                row[5] = str(mdata.sort_value)
                row[6] = mdata.data_url
                matched += 1
            else:
                # clear the cells if no data found
                row[5] = ""
                row[6] = ""

            # contact data is available here for future use, e.g. contacts.get(manufacturer)
            writer.writerow(row)

    print(f"Processed {processed} lines, matched {matched} manufacturers. Wrote {dst_csv}")


if __name__ == "__main__":
    logos = load_manufacturer_logos()
    contacts = load_manufacturer_contacts()
    print(f"Loaded {len(logos)} logos and {len(contacts)} contacts")
    update_manufacturers_csv(logos, contacts=contacts)
